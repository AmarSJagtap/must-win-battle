from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, timezone, timedelta

from database import engine, Base, get_db
from models import (
    Project, Milestone, Action, TeamMember, Review, Attachment, Reminder, User, UserPermission, Activity, KnowledgeDoc
)
from schemas import (
    ProjectCreate, ProjectUpdate, ProjectOut,
    MilestoneCreate, MilestoneOut,
    ActionCreate, ActionOut,
    TeamMemberCreate, TeamMemberOut,
    ReviewCreate, ReviewOut,
    AttachmentCreate, AttachmentOut,
    ReminderCreate, ReminderOut,
    SignUpRequest, SignInRequest, TokenOut, UserOut,
    SetPermissionsRequest, ToggleActiveRequest, ActivityOut,
    AIAssistRequest, AIAssistResponse,
    ChatRequest, ChatResponse, KnowledgeDocOut
)
from auth import (
    hash_password, verify_password, create_access_token,
    require_user, require_admin, SCREENS,
)

Base.metadata.create_all(bind=engine)

app = FastAPI(title="MWB Tracker API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

import os
from dotenv import load_dotenv
load_dotenv()

# ──── Azure OpenAI Config (set via environment variables or .env) ────
AZURE_OPENAI_ENDPOINT = os.environ.get("AZURE_OPENAI_ENDPOINT", "")
AZURE_OPENAI_KEY = os.environ.get("AZURE_OPENAI_KEY", "")
AZURE_OPENAI_DEPLOYMENT = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o-mini")
AZURE_OPENAI_API_VERSION = os.environ.get("AZURE_OPENAI_API_VERSION", "2025-01-01-preview")


@app.on_event("startup")
def startup():
    db = next(get_db())
    # seed_data(db) # Removed seeding
    # Create default admin if not exists
    if not db.query(User).filter(User.username == "admin").first():
        admin = User(
            username="admin", email="admin@mwb.com", full_name="Administrator",
            hashed_password=hash_password("admin123"), is_admin=True, is_active=True,
        )
        db.add(admin)
        db.flush()
        # Admin gets all screens
        for s in SCREENS:
            db.add(UserPermission(user_id=admin.id, screen=s, can_access=True))
        db.commit()
    db.close()


# ──── Activity Logger ────
def log_activity(db: Session, user: User, text: str, project_id: Optional[int] = None, icon: str = "info", color: str = "gray"):
    act = Activity(
        text=text,
        user_name=user.full_name or user.username,
        project_id=project_id,
        timestamp=datetime.now(timezone.utc).isoformat(),
        icon=icon,
        color=color
    )
    db.add(act)
    db.commit()


# ──── Auth ────
@app.post("/api/auth/signup", response_model=TokenOut)
def signup(data: SignUpRequest, db: Session = Depends(get_db)):
    if db.query(User).filter((User.username == data.username) | (User.email == data.email)).first():
        raise HTTPException(400, "Username or email already exists")
    user = User(
        username=data.username, email=data.email, full_name=data.full_name,
        hashed_password=hash_password(data.password), is_admin=False, is_active=True,
    )
    db.add(user)
    db.flush()
    # Log activity before committing the user creation fully
    log_activity(db, user, f"New user '{data.username}' signed up.", icon="add", color="green")
    # New users get dashboard access only by default
    db.add(UserPermission(user_id=user.id, screen="dashboard", can_access=True))
    db.commit()
    token = create_access_token({"sub": str(user.id)})
    return {"access_token": token}


@app.post("/api/auth/signin", response_model=TokenOut)
def signin(data: SignInRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == data.username).first()
    if not user or not verify_password(data.password, user.hashed_password):
        raise HTTPException(401, "Invalid credentials")
    if not user.is_active:
        raise HTTPException(403, "Account disabled")
    token = create_access_token({"sub": str(user.id)})
    return {"access_token": token}


@app.get("/api/auth/me", response_model=UserOut)
def get_me(user: User = Depends(require_user)):
    return user


# ──── Admin: User Management ────
@app.get("/api/admin/users", response_model=List[UserOut])
def list_users(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    return db.query(User).all()


@app.put("/api/admin/users/{uid}/permissions", response_model=UserOut)
def set_user_permissions(uid: int, data: SetPermissionsRequest, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).get(uid)
    if not user:
        raise HTTPException(404, "User not found")
    # Delete existing permissions
    db.query(UserPermission).filter(UserPermission.user_id == uid).delete()
    # Add new permissions
    for s in data.screens:
        if s in SCREENS:
            db.add(UserPermission(user_id=uid, screen=s, can_access=True))
    db.commit()
    db.refresh(user)
    return user


@app.put("/api/admin/users/{uid}/toggle-active", response_model=UserOut)
def toggle_user_active(uid: int, data: ToggleActiveRequest, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).get(uid)
    if not user:
        raise HTTPException(404, "User not found")
    user.is_active = data.is_active
    db.commit()
    db.refresh(user)
    return user


@app.put("/api/admin/users/{uid}/make-admin", response_model=UserOut)
def make_admin(uid: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).get(uid)
    if not user:
        raise HTTPException(404, "User not found")
    user.is_admin = True
    # Grant all screens
    db.query(UserPermission).filter(UserPermission.user_id == uid).delete()
    for s in SCREENS:
        db.add(UserPermission(user_id=uid, screen=s, can_access=True))
    db.commit()
    db.refresh(user)
    return user


# ──── Projects ────
@app.get("/api/projects", response_model=List[ProjectOut])
def list_projects(db: Session = Depends(get_db)):
    return db.query(Project).all()


@app.get("/api/projects/{pid}", response_model=ProjectOut)
def get_project(pid: int, db: Session = Depends(get_db)):
    p = db.query(Project).get(pid)
    if not p: raise HTTPException(404, "Project not found")
    return p


@app.post("/api/projects", response_model=ProjectOut)
def create_project(data: ProjectCreate, user: User = Depends(require_user), db: Session = Depends(get_db)):
    p = Project(**data.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    log_activity(db, user, f"New project '{p.title}' created.", project_id=p.id, icon="add", color="blue")
    return p


@app.put("/api/projects/{pid}", response_model=ProjectOut)
def update_project(pid: int, data: ProjectUpdate, user: User = Depends(require_user), db: Session = Depends(get_db)):
    p = db.query(Project).get(pid)
    if not p: raise HTTPException(404, "Project not found")

    old_status = p.status
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    
    db.commit()
    db.refresh(p)

    if data.status and data.status != old_status:
        log_activity(db, user, f"Project '{p.title}' status changed to {data.status}.", project_id=p.id, icon="edit", color="amber")

    return p


@app.delete("/api/projects/{pid}")
def delete_project(pid: int, user: User = Depends(require_user), db: Session = Depends(get_db)):
    p = db.query(Project).get(pid)
    if not p: raise HTTPException(404, "Project not found")
    # Log without project_id so the activity survives the cascade delete
    log_activity(db, user, f"Project '{p.title}' deleted.", project_id=None, icon="delete", color="red")
    db.delete(p)
    db.commit()
    return {"ok": True}


# ──── Milestones ────
@app.post("/api/projects/{pid}/milestones", response_model=MilestoneOut)
def create_milestone(pid: int, data: MilestoneCreate, user: User = Depends(require_user), db: Session = Depends(get_db)):
    m = Milestone(project_id=pid, **data.model_dump())
    db.add(m)
    db.commit()
    db.refresh(m)
    log_activity(db, user, f"New milestone '{m.title}' added to project.", project_id=pid, icon="add", color="purple")
    return m

@app.put("/api/milestones/{mid}", response_model=MilestoneOut)
def update_milestone(mid: int, data: MilestoneCreate, db: Session = Depends(get_db)):
    m = db.query(Milestone).get(mid)
    if not m: raise HTTPException(404, "Milestone not found")
    for k, v in data.model_dump().items(): setattr(m, k, v)
    db.commit(); db.refresh(m)
    return m

@app.delete("/api/milestones/{mid}")
def delete_milestone(mid: int, db: Session = Depends(get_db)):
    m = db.query(Milestone).get(mid)
    if not m: raise HTTPException(404, "Milestone not found")
    db.delete(m); db.commit()
    return {"ok": True}


# ──── Actions ────
@app.post("/api/projects/{pid}/actions", response_model=ActionOut)
def create_action(pid: int, data: ActionCreate, user: User = Depends(require_user), db: Session = Depends(get_db)):
    a = Action(project_id=pid, **data.model_dump())
    db.add(a)
    db.commit()
    db.refresh(a)
    log_activity(db, user, f"New action created for project.", project_id=pid, icon="add", color="gray")
    return a

@app.put("/api/actions/{aid}", response_model=ActionOut)
def update_action(aid: int, data: ActionCreate, db: Session = Depends(get_db)):
    a = db.query(Action).get(aid)
    if not a: raise HTTPException(404, "Action not found")
    for k, v in data.model_dump().items(): setattr(a, k, v)
    db.commit(); db.refresh(a)
    return a

@app.patch("/api/actions/{aid}/toggle", response_model=ActionOut)
def toggle_action(aid: int, user: User = Depends(require_user), db: Session = Depends(get_db)):
    a = db.query(Action).get(aid)
    if not a: raise HTTPException(404, "Action not found")
    a.done = not a.done
    a.status = "complete" if a.done else "open"
    db.commit()
    db.refresh(a)
    if a.done:
        log_activity(db, user, f"Action '{a.title[:30]}...' completed.", project_id=a.project_id, icon="check", color="green")
    return a

@app.delete("/api/actions/{aid}")
def delete_action(aid: int, db: Session = Depends(get_db)):
    a = db.query(Action).get(aid)
    if not a: raise HTTPException(404, "Action not found")
    db.delete(a); db.commit()
    return {"ok": True}


# ──── Team Members ────
@app.post("/api/projects/{pid}/team", response_model=TeamMemberOut)
def create_team(pid: int, data: TeamMemberCreate, db: Session = Depends(get_db)):
    t = TeamMember(project_id=pid, **data.model_dump())
    db.add(t); db.commit(); db.refresh(t)
    return t

@app.put("/api/team/{tid}", response_model=TeamMemberOut)
def update_team(tid: int, data: TeamMemberCreate, db: Session = Depends(get_db)):
    t = db.query(TeamMember).get(tid)
    if not t: raise HTTPException(404, "Team member not found")
    for k, v in data.model_dump().items(): setattr(t, k, v)
    db.commit(); db.refresh(t)
    return t

@app.delete("/api/team/{tid}")
def delete_team(tid: int, db: Session = Depends(get_db)):
    t = db.query(TeamMember).get(tid)
    if not t: raise HTTPException(404, "Team member not found")
    db.delete(t); db.commit()
    return {"ok": True}


# ──── Reviews ────
@app.post("/api/projects/{pid}/reviews", response_model=ReviewOut)
def create_review(pid: int, data: ReviewCreate, user: User = Depends(require_user), db: Session = Depends(get_db)):
    r = Review(project_id=pid, **data.model_dump())
    db.add(r)
    db.commit()
    db.refresh(r)
    # Update project status/progress from review
    p = db.query(Project).get(pid)
    if p:
        if data.status: p.status = data.status
        if data.progress is not None: p.progress = data.progress
        db.commit()
        log_activity(db, user, f"New review posted for project '{p.title}'.", project_id=pid, icon="comment", color="blue")
    return r

@app.delete("/api/reviews/{rid}")
def delete_review(rid: int, db: Session = Depends(get_db)):
    r = db.query(Review).get(rid)
    if not r: raise HTTPException(404, "Review not found")
    db.delete(r); db.commit()
    return {"ok": True}


# ──── Attachments ────
@app.post("/api/projects/{pid}/attachments", response_model=AttachmentOut)
def create_attachment(pid: int, data: AttachmentCreate, db: Session = Depends(get_db)):
    a = Attachment(project_id=pid, **data.model_dump())
    db.add(a); db.commit(); db.refresh(a)
    return a

@app.delete("/api/attachments/{aid}")
def delete_attachment(aid: int, db: Session = Depends(get_db)):
    a = db.query(Attachment).get(aid)
    if not a: raise HTTPException(404, "Attachment not found")
    db.delete(a); db.commit()
    return {"ok": True}


# ──── Reminders ────
@app.post("/api/projects/{pid}/reminders", response_model=ReminderOut)
def create_reminder(pid: int, data: ReminderCreate, db: Session = Depends(get_db)):
    r = Reminder(project_id=pid, **data.model_dump())
    db.add(r); db.commit(); db.refresh(r)
    return r

@app.delete("/api/reminders/{rid}")
def delete_reminder(rid: int, db: Session = Depends(get_db)):
    r = db.query(Reminder).get(rid)
    if not r: raise HTTPException(404, "Reminder not found")
    db.delete(r); db.commit()
    return {"ok": True}

@app.patch("/api/reminders/{rid}/complete", response_model=ReminderOut)
def complete_reminder(rid: int, db: Session = Depends(get_db)):
    r = db.query(Reminder).get(rid)
    if not r: raise HTTPException(404, "Reminder not found")
    r.completed = True
    db.commit(); db.refresh(r)
    return r

@app.patch("/api/reminders/{rid}/snooze")
def snooze_reminder(rid: int, db: Session = Depends(get_db)):
    r = db.query(Reminder).get(rid)
    if not r: raise HTTPException(404, "Reminder not found")
    # Snooze for 30 minutes
    snooze_until = datetime.now(timezone.utc) + timedelta(minutes=30)
    r.snoozed_until = snooze_until.isoformat()
    db.commit(); db.refresh(r)
    return {"ok": True, "snoozed_until": r.snoozed_until}

@app.get("/api/active-reminders", response_model=List[ReminderOut])
def get_active_reminders(db: Session = Depends(get_db)):
    """Get all reminders that are not completed and whose date is today or past, and not snoozed."""
    import re
    today = datetime.now(timezone.utc).date()
    now_iso = datetime.now(timezone.utc).isoformat()

    def parse_date(s):
        if not s: return None
        # Try YYYY-MM-DD
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s.strip())
        if m: return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        # Try "28 Apr 2026" or "28 April 2026"
        try:
            from datetime import date as dt_date
            import locale
            # Parse with strptime
            for fmt in ('%d %b %Y', '%d %B %Y', '%b %d %Y', '%B %d %Y', '%Y/%m/%d', '%m/%d/%Y'):
                try: return datetime.strptime(s.strip(), fmt).date()
                except: pass
        except: pass
        return None

    reminders = db.query(Reminder).filter(
        Reminder.completed == False,
    ).all()
    active = []
    for r in reminders:
        rem_date = parse_date(r.date)
        if rem_date is None or rem_date > today:
            continue  # Not yet due
        # Check snooze
        if r.snoozed_until and r.snoozed_until > now_iso:
            continue  # Still snoozed
        active.append(r)
    return active


# ──── Stats & Activity ────
@app.get("/api/activity", response_model=List[ActivityOut])
def get_activity(db: Session = Depends(get_db)):
    return db.query(Activity).order_by(Activity.id.desc()).limit(10).all()

@app.get("/api/upcoming-reviews")
def get_upcoming_reviews(db: Session = Depends(get_db)):
    # Get recent reviews from the reviews table
    reviews = db.query(Review).order_by(Review.id.desc()).limit(5).all()
    result = []
    for r in reviews:
        result.append({
            "id": r.id,
            "project_id": r.project_id,
            "text": r.notes[:80] if r.notes else "Status Review",
            "date": r.date,
            "author": r.author,
            "status": r.status,
            "type": "review",
        })

    # Also include upcoming reminders
    today = datetime.now(timezone.utc).date()
    reminders = db.query(Reminder).filter(Reminder.date >= str(today)).order_by(Reminder.date).limit(5).all()
    for rem in reminders:
        result.append({
            "id": rem.id + 100000,  # offset to avoid id collision
            "project_id": rem.project_id,
            "text": rem.text,
            "date": rem.date,
            "author": "",
            "status": "",
            "type": "reminder",
        })

    return result


@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    projects = db.query(Project).all()
    actions = db.query(Action).all()
    return {
        "total": len(projects),
        "on_track": sum(1 for p in projects if p.status == "On Track"),
        "at_risk": sum(1 for p in projects if p.status == "At Risk"),
        "off_track": sum(1 for p in projects if p.status == "Off Track"),
        "open_actions": sum(1 for a in actions if not a.done),
        "overdue_actions": sum(1 for a in actions if a.status == "overdue"),
    }

# ──── AI Assistant ────
# ──── Chatbot ────
def _build_db_context(db: Session) -> str:
    """Build a comprehensive text summary of the entire SQLite database for the chatbot."""
    lines = ["=== MWB Tracker Knowledge Base ==="]
    projects = db.query(Project).all()
    lines.append(f"\nTotal Projects: {len(projects)}")
    for p in projects:
        lines.append(f"\n--- Project: {p.title} ({p.proj_id}) ---")
        lines.append(f"  Status: {p.status} | Priority: {p.priority} | Progress: {p.progress}%")
        lines.append(f"  Dept: {p.dept} | Sponsor: {p.sponsor} | Lead: {p.lead}")
        lines.append(f"  Timeline: {p.start} to {p.end}")
        if p.background: lines.append(f"  Background: {p.background}")
        if p.scope:      lines.append(f"  Scope: {p.scope}")
        if p.risks:      lines.append(f"  Risks: {p.risks}")
        if p.kpis:       lines.append(f"  KPIs: {p.kpis}")
        actions = db.query(Action).filter_by(project_id=p.id).all()
        if actions:
            lines.append(f"  Actions ({len(actions)}):")
            for a in actions:
                lines.append(f"    - [{a.status.upper()}] {a.title} | Due: {a.due} | Owner: {a.responsible} | Done: {a.done}")
        milestones = db.query(Milestone).filter_by(project_id=p.id).all()
        if milestones:
            lines.append(f"  Milestones ({len(milestones)}):")
            for m in milestones:
                lines.append(f"    - [{m.status}] {m.title} | Planned: {m.planned}")
        reviews = db.query(Review).filter_by(project_id=p.id).order_by(Review.id.desc()).limit(3).all()
        if reviews:
            lines.append(f"  Recent Reviews:")
            for r in reviews:
                lines.append(f"    - {r.date} ({r.status}): {r.notes}")
        reminders = db.query(Reminder).filter_by(project_id=p.id, completed=False).all()
        if reminders:
            lines.append(f"  Active Reminders:")
            for r in reminders:
                lines.append(f"    - {r.text} | Date: {r.date}")
    return "\n".join(lines)


@app.post("/api/chat", response_model=ChatResponse)
def chat(req: ChatRequest, db: Session = Depends(get_db)):
    """Main chatbot endpoint — uses full DB context + uploaded knowledge docs."""
    db_context = _build_db_context(db)

    # Append uploaded document knowledge
    docs = db.query(KnowledgeDoc).all()
    doc_context = ""
    if docs:
        doc_context = "\n\n=== Uploaded Knowledge Documents ===\n"
        for d in docs:
            doc_context += f"\n[{d.filename}]\n{d.content[:3000]}\n"  # cap per doc

    system_prompt = (
        "You are an expert project management AI assistant for the MWB (Must-Win Battle) Tracker.\n"
        "You have access to the full knowledge base below which includes all projects, actions, milestones, "
        "reviews, reminders, and any uploaded documents.\n"
        "Answer the user's question accurately and helpfully. "
        "If a project is Off Track or At Risk, suggest concrete improvement strategies. "
        "Be concise but thorough. Use bullet points where helpful.\n\n"
        f"{db_context}{doc_context}"
    )

    try:
        from openai import AzureOpenAI
        client = AzureOpenAI(
            azure_endpoint=AZURE_OPENAI_ENDPOINT,
            api_key=AZURE_OPENAI_KEY,
            api_version=AZURE_OPENAI_API_VERSION,
        )
        response = client.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": req.message}
            ],
            max_tokens=1500,
        )
        return {"reply": response.choices[0].message.content}
    except Exception as e:
        return {"reply": f"⚠️ Error: {str(e)}"}


@app.post("/api/knowledge-docs/upload", response_model=KnowledgeDocOut)
async def upload_knowledge_doc(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload a PDF or Excel file — extract text and store in knowledge base."""
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower()
    raw = await file.read()
    content = ""

    if ext == "pdf":
        try:
            import io
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(raw))
            content = "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception as e:
            content = f"[PDF parse error: {e}]"
    elif ext in ("xlsx", "xls", "xlsm", "csv"):
        try:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows.append(f"== Sheet: {sheet} ==")
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
            content = "\n".join(rows)
        except Exception as e:
            content = f"[Excel parse error: {e}]"
    else:
        raise HTTPException(400, "Only PDF and Excel files are supported")

    doc = KnowledgeDoc(
        filename=filename,
        filetype=ext,
        content=content[:50000],  # cap at 50k chars
        uploaded_at=datetime.now(timezone.utc).isoformat()
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return doc


@app.get("/api/knowledge-docs", response_model=List[KnowledgeDocOut])
def list_knowledge_docs(db: Session = Depends(get_db)):
    return db.query(KnowledgeDoc).order_by(KnowledgeDoc.id.desc()).all()


@app.delete("/api/knowledge-docs/{doc_id}")
def delete_knowledge_doc(doc_id: int, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDoc).get(doc_id)
    if not doc:
        raise HTTPException(404, "Document not found")
    db.delete(doc)
    db.commit()
    return {"ok": True}
